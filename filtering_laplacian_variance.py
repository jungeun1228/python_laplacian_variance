#convert csv to excel
import pandas as pd
import csv
import os
import openpyxl
from openpyxl import Workbook
wb = openpyxl.Workbook()
ws = wb.active

#update path 2 / 2_2/  3  
path1 = "D:/Images/folder_name"
path2 = "image_analysis_file_name.csv"
path2_2 = "image_analysis_file_name.xlsx"
path3 = "main_folder_name"
path4 = "TimePoint_1/"
csvpath = os.path.join(path1,path2)
excelpath = os.path.join(path1,path2_2)
imagepath = os.path.join(path1,path3,path4)
with open(csvpath) as f:
    reader = csv.reader(f, delimiter=',')
    for row in reader:
        ws.append(row)

wb.save(excelpath)

#organize images
import numpy as np
import cv2
from pathlib import Path
main_folder = Path(imagepath)
#open excel file
excelfile = excelpath

# group images for wavelength 1 (nuclei image)
# & wavelength 2 (mmpi - )
# & wavelength 3 (mmpi - )
# & wavelength 4 (phase-contrast image)
from natsort import natsorted
filelistw1 = natsorted(main_folder.glob("*w1.TIF"))
filelistw2 = natsorted(main_folder.glob("*w2.TIF"))
filelistw3 = natsorted(main_folder.glob("*w3.TIF"))
filelistw4 = natsorted(main_folder.glob("*w4.TIF"))

listw1 = []
listw2 = []
listw3 = []
listw4 = []

# (1) Read the image
# (2) Calculate the laplacian variance of this image 
# (3) Create a list
for i in filelistw1:
    imagepath1 = os.path.join(main_folder,i)
    imgw1 = cv2.imread(imagepath1)
    var1 = cv2.Laplacian(imgw1, cv2.CV_64F).var()
    listw1.append(var1)
for i in filelistw2:
    imagepath2 = os.path.join(main_folder,i)
    imgw2 = cv2.imread(imagepath2)
    var2 = cv2.Laplacian(imgw2, cv2.CV_64F).var()
    listw2.append(var2)
for i in filelistw3:
    imagepath3 = os.path.join(main_folder,i)
    imgw3 = cv2.imread(imagepath3)
    var3 = cv2.Laplacian(imgw3, cv2.CV_64F).var()
    listw3.append(var3)
for i in filelistw4:
    imagepath4 = os.path.join(main_folder,i)
    imgw4 = cv2.imread(imagepath4)
    gray4 = cv2.cvtColor(imgw4, cv2.COLOR_BGR2GRAY)
    var4 = cv2.Laplacian(gray4, cv2.CV_64F).var()    
    listw4.append(var4)
    
#add variance to excel file
AllLaplacian1 = pd.DataFrame([listw1, listw2, listw3, listw4])
AllLaplacian2 = AllLaplacian1.transpose()
AllLaplacian2 = AllLaplacian2.rename({0:"Laplacian variance_w1",1:"Laplacian variance_w2",2:"Laplacian variance_w3",3:"Laplacian variance_w4"}, axis="columns")

from openpyxl import load_workbook
workbook = load_workbook(excelfile)
writer = pd.ExcelWriter(excelfile, engine='openpyxl') 
writer.book = workbook

AllLaplacian2.to_excel(writer, index=False, sheet_name="Sheet", startcol=15)

#writer.save()
writer.close()


#Bring Red/green ratio & cellconfluency
OrigCellconf = pd.read_excel(excelfile, usecols = 'I')
OrigRGratio = pd.read_excel(excelfile, usecols = 'J')

#open excel file
excelfile = "D:/Images/Practice/image_analysis_excel_file.xlsx"


#set threshold
threshold = 10
#filter cellconfluence compared to threshold
for i in excelfile.index :
       if excelfile.iloc[i, 15] < threshold and excelfile.iloc[i, 18] < threshold :
                excelfile.iloc[i, 19] = excelfile.iat[i, 8]
                     
#filter RGratio compared to threshold
for i in excelfile.index :
    if excelfile.iloc[i, 16] < threshold and excelfile.iloc[i, 17] < threshold :
                excelfile.iloc[i, 20] = excelfile.iat[i, 9]
                     
